# Backend/tests/test_excel_generator.py

import pytest
from openpyxl import load_workbook
import io

from app.utils.excel_generator import generate_excel


class TestExcelGenerationBasic:
    """Test basic Excel generation functionality."""
    
    def test_generate_excel_financial_statement(self):
        """Test Excel generation for financial statement template."""
        data = {
            "statements": [
                {
                    "statement_type": "balance_sheet",
                    "period_start": "2024-01-01T00:00:00",
                    "period_end": "2024-12-31T23:59:59",
                    "metrics": {
                        "gross_margin": 35.5,
                        "net_margin": 18.2,
                        "roa": 12.5,
                        "roe": 15.8
                    }
                }
            ]
        }
        
        result = generate_excel("financial_statement", data)
        
        assert result is not None
        assert isinstance(result, bytes)
        assert len(result) > 0
        
        # Verify it's a valid Excel file
        wb = load_workbook(io.BytesIO(result))
        assert "Financial Statements" in wb.sheetnames
    
    def test_generate_excel_costing_analysis(self):
        """Test Excel generation for costing analysis template."""
        data = {
            "products": [
                {
                    "name": "Product A",
                    "category": "Electronics",
                    "costing": {
                        "raw_material_cost": 100.0,
                        "labour_cost_per_unit": 20.0,
                        "overhead_percentage": 10.0,
                        "target_margin_percentage": 25.0
                    },
                    "statistics": {
                        "avg_margin": 25.5
                    }
                }
            ]
        }
        
        result = generate_excel("costing_analysis", data)
        
        assert result is not None
        assert isinstance(result, bytes)
        assert len(result) > 0
        
        wb = load_workbook(io.BytesIO(result))
        assert "Costing Analysis" in wb.sheetnames
    
    def test_generate_excel_order_evaluation(self):
        """Test Excel generation for order evaluation template."""
        data = {
            "orders": [
                {
                    "order_number": "ORD-001",
                    "customer_name": "Customer A",
                    "order_date": "2024-01-15T10:30:00",
                    "financials": {
                        "total_selling_price": 5000.0,
                        "total_cost": 3750.0,
                        "margin_percentage": 25.0
                    },
                    "evaluation": {
                        "profitability_score": 85,
                        "risk_level": "Low",
                        "should_accept": True
                    }
                }
            ]
        }
        
        result = generate_excel("order_evaluation", data)
        
        assert result is not None
        assert isinstance(result, bytes)
        
        wb = load_workbook(io.BytesIO(result))
        assert "Order Evaluations" in wb.sheetnames
    
    def test_generate_excel_margin_analysis(self):
        """Test Excel generation for margin analysis template."""
        data = {
            "groups": [
                {
                    "name": "Product A",
                    "total_revenue": 10000.0,
                    "total_cost": 7500.0,
                    "margin": 2500.0,
                    "margin_percentage": 25.0,
                    "order_count": 5
                }
            ]
        }
        
        result = generate_excel("margin_analysis", data)
        
        assert result is not None
        assert isinstance(result, bytes)
        
        wb = load_workbook(io.BytesIO(result))
        assert "Margin Analysis" in wb.sheetnames
    
    def test_generate_excel_receivables_report(self):
        """Test Excel generation for receivables report template."""
        data = {
            "aging_summary": {
                "0-30 days": {"count": 5, "total_amount": 20000.0},
                "31-60 days": {"count": 7, "total_amount": 20000.0},
                "61-90 days": {"count": 3, "total_amount": 10000.0}
            },
            "receivables": [
                {
                    "party_name": "Customer A",
                    "amount": 5000.0,
                    "transaction_date": "2024-11-15T00:00:00",
                    "due_date": "2024-12-15T00:00:00",
                    "days_outstanding": 30,
                    "is_overdue": False
                }
            ]
        }
        
        result = generate_excel("receivables_report", data)
        
        assert result is not None
        assert isinstance(result, bytes)
        
        wb = load_workbook(io.BytesIO(result))
        assert "Summary" in wb.sheetnames
        assert "Details" in wb.sheetnames
    
    def test_generate_excel_unknown_template(self):
        """Test Excel generation with unknown template raises error."""
        data = {"test": "data"}
        
        with pytest.raises(ValueError) as exc_info:
            generate_excel("unknown_template", data)
        
        assert "Unknown template" in str(exc_info.value)


class TestExcelMultipleWorksheets:
    """Test Excel generation with multiple worksheets."""
    
    def test_receivables_report_has_two_worksheets(self):
        """Test that receivables report creates Summary and Details worksheets."""
        data = {
            "aging_summary": {
                "0-30 days": {"count": 5, "total_amount": 20000.0}
            },
            "receivables": [
                {
                    "party_name": "Customer A",
                    "amount": 5000.0,
                    "transaction_date": "2024-11-15T00:00:00",
                    "due_date": "2024-12-15T00:00:00",
                    "days_outstanding": 30,
                    "is_overdue": False
                }
            ]
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        
        assert len(wb.sheetnames) == 2
        assert "Summary" in wb.sheetnames
        assert "Details" in wb.sheetnames
    
    def test_receivables_summary_worksheet_content(self):
        """Test that Summary worksheet contains aging bucket data."""
        data = {
            "aging_summary": {
                "0-30 days": {"count": 5, "total_amount": 20000.0},
                "31-60 days": {"count": 3, "total_amount": 15000.0}
            },
            "receivables": []
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Summary"]
        
        # Check headers
        assert ws.cell(row=1, column=1).value == "Aging Bucket"
        assert ws.cell(row=1, column=2).value == "Count"
        assert ws.cell(row=1, column=3).value == "Amount"
        
        # Check data rows
        assert ws.cell(row=2, column=1).value == "0-30 days"
        assert ws.cell(row=2, column=2).value == 5
        assert ws.cell(row=2, column=3).value == 20000.0
    
    def test_receivables_details_worksheet_content(self):
        """Test that Details worksheet contains receivable records."""
        data = {
            "aging_summary": {},
            "receivables": [
                {
                    "party_name": "Customer A",
                    "amount": 5000.0,
                    "transaction_date": "2024-11-15T00:00:00",
                    "due_date": "2024-12-15T00:00:00",
                    "days_outstanding": 30,
                    "is_overdue": False
                }
            ]
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Details"]
        
        # Check headers
        assert ws.cell(row=1, column=1).value == "Party Name"
        assert ws.cell(row=1, column=2).value == "Amount"
        
        # Check data
        assert ws.cell(row=2, column=1).value == "Customer A"
        assert ws.cell(row=2, column=2).value == 5000.0
        assert ws.cell(row=2, column=6).value == "No"


class TestExcelDataFormatting:
    """Test Excel data formatting."""
    
    def test_financial_statement_header_formatting(self):
        """Test that headers are formatted with bold and background color."""
        data = {
            "statements": []
        }
        
        result = generate_excel("financial_statement", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Financial Statements"]
        
        # Check first header cell
        header_cell = ws.cell(row=1, column=1)
        assert header_cell.font.bold is True
        # Color can be FFFFFFFF or 00FFFFFF depending on openpyxl version
        assert header_cell.font.color.rgb in ["FFFFFFFF", "00FFFFFF"]
        assert header_cell.fill.start_color.rgb in ["FF366092", "00366092"]
    
    def test_column_width_auto_adjustment(self):
        """Test that column widths are auto-adjusted based on content."""
        data = {
            "products": [
                {
                    "name": "Very Long Product Name That Should Adjust Column Width",
                    "category": "Electronics",
                    "costing": {
                        "raw_material_cost": 100.0,
                        "labour_cost_per_unit": 20.0,
                        "overhead_percentage": 10.0,
                        "target_margin_percentage": 25.0
                    },
                    "statistics": {
                        "avg_margin": 25.5
                    }
                }
            ]
        }
        
        result = generate_excel("costing_analysis", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Costing Analysis"]
        
        # Column width should be adjusted (but capped at 50)
        assert ws.column_dimensions['A'].width > 10
        assert ws.column_dimensions['A'].width <= 50
    
    def test_date_formatting(self):
        """Test that dates are formatted correctly (YYYY-MM-DD)."""
        data = {
            "statements": [
                {
                    "statement_type": "balance_sheet",
                    "period_start": "2024-01-15T10:30:45",
                    "period_end": "2024-12-31T23:59:59",
                    "metrics": {
                        "gross_margin": 35.5,
                        "net_margin": 18.2,
                        "roa": 12.5,
                        "roe": 15.8
                    }
                }
            ]
        }
        
        result = generate_excel("financial_statement", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Financial Statements"]
        
        # Dates should be truncated to YYYY-MM-DD
        assert ws.cell(row=2, column=2).value == "2024-01-15"
        assert ws.cell(row=2, column=3).value == "2024-12-31"
    
    def test_numeric_values_preserved(self):
        """Test that numeric values are preserved as numbers, not strings."""
        data = {
            "groups": [
                {
                    "name": "Product A",
                    "total_revenue": 10000.50,
                    "total_cost": 7500.25,
                    "margin": 2500.25,
                    "margin_percentage": 25.0025,
                    "order_count": 5
                }
            ]
        }
        
        result = generate_excel("margin_analysis", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Margin Analysis"]
        
        # Check that numeric values are numbers
        assert isinstance(ws.cell(row=2, column=2).value, (int, float))
        assert ws.cell(row=2, column=2).value == 10000.50
        assert isinstance(ws.cell(row=2, column=6).value, int)
        assert ws.cell(row=2, column=6).value == 5
    
    def test_boolean_to_text_conversion(self):
        """Test that boolean values are converted to Accept/Review text."""
        data = {
            "orders": [
                {
                    "order_number": "ORD-001",
                    "customer_name": "Customer A",
                    "order_date": "2024-01-15T10:30:00",
                    "financials": {
                        "total_selling_price": 5000.0,
                        "total_cost": 3750.0,
                        "margin_percentage": 25.0
                    },
                    "evaluation": {
                        "profitability_score": 85,
                        "risk_level": "Low",
                        "should_accept": True
                    }
                },
                {
                    "order_number": "ORD-002",
                    "customer_name": "Customer B",
                    "order_date": "2024-01-16T10:30:00",
                    "financials": {
                        "total_selling_price": 3000.0,
                        "total_cost": 2900.0,
                        "margin_percentage": 3.3
                    },
                    "evaluation": {
                        "profitability_score": 45,
                        "risk_level": "High",
                        "should_accept": False
                    }
                }
            ]
        }
        
        result = generate_excel("order_evaluation", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Order Evaluations"]
        
        assert ws.cell(row=2, column=9).value == "Accept"
        assert ws.cell(row=3, column=9).value == "Review"


class TestExcelSpecialCharacters:
    """Test Excel generation with special characters."""
    
    def test_product_name_with_special_characters(self):
        """Test that product names with special characters are handled correctly."""
        data = {
            "products": [
                {
                    "name": "Product & Co. (Type #1)",
                    "category": "Electronics & Gadgets",
                    "costing": {
                        "raw_material_cost": 100.0,
                        "labour_cost_per_unit": 20.0,
                        "overhead_percentage": 10.0,
                        "target_margin_percentage": 25.0
                    },
                    "statistics": {
                        "avg_margin": 25.5
                    }
                }
            ]
        }
        
        result = generate_excel("costing_analysis", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Costing Analysis"]
        
        assert ws.cell(row=2, column=1).value == "Product & Co. (Type #1)"
        assert ws.cell(row=2, column=2).value == "Electronics & Gadgets"
    
    def test_customer_name_with_unicode(self):
        """Test that customer names with Unicode characters are handled."""
        data = {
            "orders": [
                {
                    "order_number": "ORD-001",
                    "customer_name": "Café René & Søn",
                    "order_date": "2024-01-15T10:30:00",
                    "financials": {
                        "total_selling_price": 5000.0,
                        "total_cost": 3750.0,
                        "margin_percentage": 25.0
                    },
                    "evaluation": {
                        "profitability_score": 85,
                        "risk_level": "Low",
                        "should_accept": True
                    }
                }
            ]
        }
        
        result = generate_excel("order_evaluation", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Order Evaluations"]
        
        assert ws.cell(row=2, column=2).value == "Café René & Søn"
    
    def test_party_name_with_quotes(self):
        """Test that party names with quotes are handled correctly."""
        data = {
            "aging_summary": {},
            "receivables": [
                {
                    "party_name": 'Company "Best" Ltd.',
                    "amount": 5000.0,
                    "transaction_date": "2024-11-15T00:00:00",
                    "due_date": "2024-12-15T00:00:00",
                    "days_outstanding": 30,
                    "is_overdue": False
                }
            ]
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Details"]
        
        assert ws.cell(row=2, column=1).value == 'Company "Best" Ltd.'


class TestExcelWithEmptyData:
    """Test Excel generation with empty or minimal data."""
    
    def test_financial_statement_empty_statements(self):
        """Test financial statement with no statements."""
        data = {
            "statements": []
        }
        
        result = generate_excel("financial_statement", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Financial Statements"]
        
        # Should have headers but no data rows
        assert ws.cell(row=1, column=1).value == "Type"
        assert ws.cell(row=2, column=1).value is None
    
    def test_costing_analysis_empty_products(self):
        """Test costing analysis with no products."""
        data = {
            "products": []
        }
        
        result = generate_excel("costing_analysis", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Costing Analysis"]
        
        # Should have headers but no data rows
        assert ws.cell(row=1, column=1).value == "Product"
        assert ws.cell(row=2, column=1).value is None
    
    def test_receivables_empty_aging_summary(self):
        """Test receivables report with empty aging summary."""
        data = {
            "aging_summary": {},
            "receivables": []
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        ws_summary = wb["Summary"]
        ws_details = wb["Details"]
        
        # Should have headers but no data
        assert ws_summary.cell(row=1, column=1).value == "Aging Bucket"
        assert ws_summary.cell(row=2, column=1).value is None
        assert ws_details.cell(row=1, column=1).value == "Party Name"
        assert ws_details.cell(row=2, column=1).value is None


class TestExcelWithMissingFields:
    """Test Excel generation with missing or None fields."""
    
    def test_missing_metrics_fields(self):
        """Test that missing metrics fields default to 0."""
        data = {
            "statements": [
                {
                    "statement_type": "balance_sheet",
                    "period_start": "2024-01-01T00:00:00",
                    "period_end": "2024-12-31T23:59:59",
                    "metrics": {}  # Empty metrics
                }
            ]
        }
        
        result = generate_excel("financial_statement", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Financial Statements"]
        
        # Should use 0 for missing metrics
        assert ws.cell(row=2, column=4).value == 0
        assert ws.cell(row=2, column=5).value == 0
    
    def test_missing_due_date(self):
        """Test that missing due_date is handled gracefully."""
        data = {
            "aging_summary": {},
            "receivables": [
                {
                    "party_name": "Customer A",
                    "amount": 5000.0,
                    "transaction_date": "2024-11-15T00:00:00",
                    "due_date": None,  # Missing due date
                    "days_outstanding": 30,
                    "is_overdue": False
                }
            ]
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Details"]
        
        # Should handle None due_date (empty string or None)
        assert ws.cell(row=2, column=4).value in ["", None]
    
    def test_missing_nested_fields(self):
        """Test that missing nested fields are handled gracefully."""
        data = {
            "products": [
                {
                    "name": "Product A",
                    "category": "Electronics",
                    "costing": {},  # Empty costing
                    "statistics": {}  # Empty statistics
                }
            ]
        }
        
        result = generate_excel("costing_analysis", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Costing Analysis"]
        
        # Should use 0 for missing fields
        assert ws.cell(row=2, column=3).value == 0
        assert ws.cell(row=2, column=7).value == 0


class TestExcelWithLargeDatasets:
    """Test Excel generation with large datasets."""
    
    def test_many_products(self):
        """Test costing analysis with many products."""
        products = []
        for i in range(100):
            products.append({
                "name": f"Product {i}",
                "category": "Electronics",
                "costing": {
                    "raw_material_cost": 100.0 + i,
                    "labour_cost_per_unit": 20.0 + i,
                    "overhead_percentage": 10.0,
                    "target_margin_percentage": 25.0
                },
                "statistics": {
                    "avg_margin": 25.0 + i * 0.1
                }
            })
        
        data = {"products": products}
        
        result = generate_excel("costing_analysis", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Costing Analysis"]
        
        # Should have all 100 products plus header row
        assert ws.max_row == 101
        assert ws.cell(row=101, column=1).value == "Product 99"
    
    def test_many_receivables(self):
        """Test receivables report with many receivables."""
        receivables = []
        for i in range(200):
            receivables.append({
                "party_name": f"Customer {i}",
                "amount": 1000.0 + i * 10,
                "transaction_date": "2024-11-15T00:00:00",
                "due_date": "2024-12-15T00:00:00",
                "days_outstanding": 30 + i,
                "is_overdue": i % 2 == 0
            })
        
        data = {
            "aging_summary": {},
            "receivables": receivables
        }
        
        result = generate_excel("receivables_report", data)
        wb = load_workbook(io.BytesIO(result))
        ws = wb["Details"]
        
        # Should have all 200 receivables plus header row
        assert ws.max_row == 201
        assert ws.cell(row=201, column=1).value == "Customer 199"


class TestExcelErrorHandling:
    """Test error handling in Excel generation."""
    
    def test_generate_excel_with_invalid_template(self):
        """Test that invalid template raises ValueError."""
        data = {"test": "data"}
        
        with pytest.raises(ValueError) as exc_info:
            generate_excel("invalid_template", data)
        
        assert "Unknown template" in str(exc_info.value)
    
    def test_generate_excel_with_none_data(self):
        """Test that None data is handled gracefully."""
        # This should raise an error or handle gracefully
        try:
            result = generate_excel("financial_statement", None)
            # If it doesn't raise, it should return valid bytes
            assert isinstance(result, bytes)
        except (TypeError, AttributeError):
            # Expected error for None data
            pass
