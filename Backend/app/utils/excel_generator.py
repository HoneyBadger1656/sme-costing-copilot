# backend/app/utils/excel_generator.py

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import Dict, Any
import io

from app.logging_config import get_logger

logger = get_logger(__name__)


def generate_excel(template_id: str, data: Dict[str, Any], options: Dict[str, Any] = None) -> bytes:
    """
    Generate Excel report from template and data.
    
    Args:
        template_id: Report template ID
        data: Report data
        options: Generation options
    
    Returns:
        Excel file as bytes
    """
    try:
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Generate worksheets based on template
        if template_id == "financial_statement":
            _generate_financial_statement_excel(wb, data)
        elif template_id == "costing_analysis":
            _generate_costing_analysis_excel(wb, data)
        elif template_id == "order_evaluation":
            _generate_order_evaluation_excel(wb, data)
        elif template_id == "margin_analysis":
            _generate_margin_analysis_excel(wb, data)
        elif template_id == "receivables_report":
            _generate_receivables_report_excel(wb, data)
        else:
            raise ValueError(f"Unknown template: {template_id}")
        
        # Save to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_bytes = buffer.getvalue()
        buffer.close()
        
        logger.info("excel_generated", template_id=template_id, size=len(excel_bytes))
        return excel_bytes
        
    except Exception as e:
        logger.error("excel_generation_error", error=str(e), template_id=template_id)
        raise


def _apply_header_style(cell):
    """Apply header styling to a cell"""
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _generate_financial_statement_excel(wb: Workbook, data: Dict[str, Any]):
    """Generate financial statement Excel worksheets"""
    ws = wb.create_sheet("Financial Statements")
    
    # Headers
    headers = ["Type", "Period Start", "Period End", "Gross Margin %", "Net Margin %", "ROA %", "ROE %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)
    
    # Data
    statements = data.get("statements", [])
    for row, stmt in enumerate(statements, 2):
        metrics = stmt.get("metrics", {})
        ws.cell(row=row, column=1, value=stmt.get("statement_type", ""))
        ws.cell(row=row, column=2, value=stmt.get("period_start", "")[:10])
        ws.cell(row=row, column=3, value=stmt.get("period_end", "")[:10])
        ws.cell(row=row, column=4, value=metrics.get("gross_margin", 0))
        ws.cell(row=row, column=5, value=metrics.get("net_margin", 0))
        ws.cell(row=row, column=6, value=metrics.get("roa", 0))
        ws.cell(row=row, column=7, value=metrics.get("roe", 0))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def _generate_costing_analysis_excel(wb: Workbook, data: Dict[str, Any]):
    """Generate costing analysis Excel worksheets"""
    ws = wb.create_sheet("Costing Analysis")
    
    headers = ["Product", "Category", "Raw Material", "Labour", "Overhead %", "Target Margin %", "Avg Margin %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)
    
    products = data.get("products", [])
    for row, product in enumerate(products, 2):
        costing = product.get("costing", {})
        stats = product.get("statistics", {})
        ws.cell(row=row, column=1, value=product.get("name", ""))
        ws.cell(row=row, column=2, value=product.get("category", ""))
        ws.cell(row=row, column=3, value=costing.get("raw_material_cost", 0))
        ws.cell(row=row, column=4, value=costing.get("labour_cost_per_unit", 0))
        ws.cell(row=row, column=5, value=costing.get("overhead_percentage", 0))
        ws.cell(row=row, column=6, value=costing.get("target_margin_percentage", 0))
        ws.cell(row=row, column=7, value=stats.get("avg_margin", 0))
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def _generate_order_evaluation_excel(wb: Workbook, data: Dict[str, Any]):
    """Generate order evaluation Excel worksheets"""
    ws = wb.create_sheet("Order Evaluations")
    
    headers = ["Order #", "Customer", "Date", "Revenue", "Cost", "Margin %", "Score", "Risk", "Recommendation"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)
    
    orders = data.get("orders", [])
    for row, order in enumerate(orders, 2):
        financials = order.get("financials", {})
        evaluation = order.get("evaluation", {})
        ws.cell(row=row, column=1, value=order.get("order_number", ""))
        ws.cell(row=row, column=2, value=order.get("customer_name", ""))
        ws.cell(row=row, column=3, value=order.get("order_date", "")[:10])
        ws.cell(row=row, column=4, value=financials.get("total_selling_price", 0))
        ws.cell(row=row, column=5, value=financials.get("total_cost", 0))
        ws.cell(row=row, column=6, value=financials.get("margin_percentage", 0))
        ws.cell(row=row, column=7, value=evaluation.get("profitability_score", 0))
        ws.cell(row=row, column=8, value=evaluation.get("risk_level", ""))
        ws.cell(row=row, column=9, value="Accept" if evaluation.get("should_accept") else "Review")
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def _generate_margin_analysis_excel(wb: Workbook, data: Dict[str, Any]):
    """Generate margin analysis Excel worksheets"""
    ws = wb.create_sheet("Margin Analysis")
    
    headers = ["Name", "Revenue", "Cost", "Margin", "Margin %", "Order Count"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        _apply_header_style(cell)
    
    groups = data.get("groups", [])
    for row, group in enumerate(groups, 2):
        ws.cell(row=row, column=1, value=group.get("name", ""))
        ws.cell(row=row, column=2, value=group.get("total_revenue", 0))
        ws.cell(row=row, column=3, value=group.get("total_cost", 0))
        ws.cell(row=row, column=4, value=group.get("margin", 0))
        ws.cell(row=row, column=5, value=group.get("margin_percentage", 0))
        ws.cell(row=row, column=6, value=group.get("order_count", 0))
    
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def _generate_receivables_report_excel(wb: Workbook, data: Dict[str, Any]):
    """Generate receivables report Excel worksheets"""
    # Summary sheet
    ws_summary = wb.create_sheet("Summary")
    ws_summary.cell(row=1, column=1, value="Aging Bucket")
    ws_summary.cell(row=1, column=2, value="Count")
    ws_summary.cell(row=1, column=3, value="Amount")
    _apply_header_style(ws_summary.cell(row=1, column=1))
    _apply_header_style(ws_summary.cell(row=1, column=2))
    _apply_header_style(ws_summary.cell(row=1, column=3))
    
    aging_summary = data.get("aging_summary", {})
    row = 2
    for bucket, info in aging_summary.items():
        ws_summary.cell(row=row, column=1, value=bucket)
        ws_summary.cell(row=row, column=2, value=info.get("count", 0))
        ws_summary.cell(row=row, column=3, value=info.get("total_amount", 0))
        row += 1
    
    # Details sheet
    ws_details = wb.create_sheet("Details")
    headers = ["Party Name", "Amount", "Transaction Date", "Due Date", "Days Outstanding", "Overdue"]
    for col, header in enumerate(headers, 1):
        cell = ws_details.cell(row=1, column=col, value=header)
        _apply_header_style(cell)
    
    receivables = data.get("receivables", [])
    for row, receivable in enumerate(receivables, 2):
        ws_details.cell(row=row, column=1, value=receivable.get("party_name", ""))
        ws_details.cell(row=row, column=2, value=receivable.get("amount", 0))
        ws_details.cell(row=row, column=3, value=receivable.get("transaction_date", "")[:10])
        ws_details.cell(row=row, column=4, value=receivable.get("due_date", "")[:10] if receivable.get("due_date") else "")
        ws_details.cell(row=row, column=5, value=receivable.get("days_outstanding", 0))
        ws_details.cell(row=row, column=6, value="Yes" if receivable.get("is_overdue") else "No")
    
    for ws in [ws_summary, ws_details]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
